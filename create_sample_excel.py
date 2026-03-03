"""
create_sample_excel.py
----------------------
Run this script once to generate a sample courses.xlsx file.
Replace the sample data with your real course information.

Usage:
    python create_sample_excel.py
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_sample_courses_excel():
    sample_data = [
        {
            "course_name": "Access to Higher Education (Science)",
            "description": "A pathway qualification for adults wanting to progress to university-level science, nursing, or healthcare degrees.",
            "level": "Level 3",
            "duration": "1 Year (Full-Time)",
            "start_date": "September 2025",
            "mode": "Blended Learning",
            "department": "Science & Health",
            "price": 1500,
            "url": "https://southlondoncollege.ac.uk/courses/access-to-he-science",
        },
        {
            "course_name": "BTEC Level 3 Extended Diploma in Business",
            "description": "A comprehensive qualification equivalent to 3 A-Levels covering marketing, finance, HR, and enterprise.",
            "level": "Level 3",
            "duration": "2 Years (Full-Time)",
            "start_date": "September 2025",
            "mode": "On-Campus",
            "department": "Business",
            "price": 0,
            "url": "https://southlondoncollege.ac.uk/courses/btec-business",
        },
        {
            "course_name": "AAT Foundation Certificate in Accounting",
            "description": "Ideal for beginners wanting to start a career in finance or accounting. Covers bookkeeping, costing, and spreadsheets.",
            "level": "Level 2",
            "duration": "6 Months (Part-Time)",
            "start_date": "January 2025 / September 2025",
            "mode": "Evening Classes",
            "department": "Finance & Accounting",
            "price": 850,
            "url": "https://southlondoncollege.ac.uk/courses/aat-foundation",
        },
        {
            "course_name": "T-Level in Digital Production, Design and Development",
            "description": "Industry-focused qualification with a 45-day industry placement. Covers software development, networking, and UX design.",
            "level": "Level 3",
            "duration": "2 Years (Full-Time)",
            "start_date": "September 2025",
            "mode": "On-Campus",
            "department": "Digital & IT",
            "price": 0,
            "url": "https://southlondoncollege.ac.uk/courses/t-level-digital",
        },
        {
            "course_name": "English GCSE Resit",
            "description": "For learners aged 16+ who need to improve or retake their English GCSE. Builds reading, writing, and communication skills.",
            "level": "Level 2 (GCSE equivalent)",
            "duration": "1 Year (Full-Time or Part-Time)",
            "start_date": "September 2025",
            "mode": "On-Campus / Online",
            "department": "English & Maths",
            "price": 0,
            "url": "https://southlondoncollege.ac.uk/courses/english-gcse-resit",
        },
        {
            "course_name": "Maths GCSE Resit",
            "description": "For learners who need to retake or improve their Maths GCSE grade. Covers number, algebra, geometry, and statistics.",
            "level": "Level 2 (GCSE equivalent)",
            "duration": "1 Year (Full-Time or Part-Time)",
            "start_date": "September 2025",
            "mode": "On-Campus / Online",
            "department": "English & Maths",
            "price": 0,
            "url": "https://southlondoncollege.ac.uk/courses/maths-gcse-resit",
        },
        {
            "course_name": "City & Guilds Level 2 Diploma in Hairdressing",
            "description": "Practical and theoretical training in cutting, colouring, styling, and salon management. Includes live client work.",
            "level": "Level 2",
            "duration": "1 Year (Full-Time)",
            "start_date": "September 2025",
            "mode": "On-Campus (Salon)",
            "department": "Hair & Beauty",
            "price": 600,
            "url": "https://southlondoncollege.ac.uk/courses/hairdressing-l2",
        },
        {
            "course_name": "HNC Construction and the Built Environment",
            "description": "Higher National Certificate for those pursuing a career in construction management, surveying, or civil engineering.",
            "level": "Level 4",
            "duration": "1 Year (Full-Time) / 2 Years (Part-Time)",
            "start_date": "September 2025",
            "mode": "Blended Learning",
            "department": "Construction",
            "price": 6000,
            "url": "https://southlondoncollege.ac.uk/courses/hnc-construction",
        },
        {
            "course_name": "Certificate in Education and Training (CET)",
            "description": "For those wanting to teach in further education or workplace settings. Includes observed teaching practice.",
            "level": "Level 4",
            "duration": "1 Year (Part-Time)",
            "start_date": "October 2025",
            "mode": "Blended Learning",
            "department": "Teacher Training",
            "price": 1200,
            "url": "https://southlondoncollege.ac.uk/courses/cet-teacher-training",
        },
        {
            "course_name": "ESOL Entry Level 3",
            "description": "English for Speakers of Other Languages. Develops everyday speaking, listening, reading, and writing skills.",
            "level": "Entry Level 3",
            "duration": "1 Year",
            "start_date": "September 2025 / January 2026",
            "mode": "On-Campus",
            "department": "ESOL",
            "price": 0,
            "url": "https://southlondoncollege.ac.uk/courses/esol-entry-3",
        },
    ]

    df = pd.DataFrame(sample_data)
    output_path = "courses.xlsx"
    df.to_excel(output_path, index=False, sheet_name="Courses")

    # --- Apply formatting ---
    wb = load_workbook(output_path)
    ws = wb["Courses"]

    header_fill = PatternFill("solid", start_color="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
    alt_fill = PatternFill("solid", start_color="EBF0F7")

    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(name="Arial", size=10)

    # Hide the price column (column I = index 9) so bot can't see raw price
    # but URL remains available
    col_widths = {
        "A": 40, "B": 55, "C": 15, "D": 28, "E": 30,
        "F": 22, "G": 22, "H": 10, "I": 45,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"✅ Sample course file created: {output_path}")
    print(f"   {len(df)} courses added across {df['department'].nunique()} departments.")
    print("\n📝 Next steps:")
    print("   1. Open courses.xlsx and replace the sample data with your real courses.")
    print("   2. Keep the same column headers (or update excel_loader.py if you rename them).")
    print("   3. Make sure every course has a 'url' column with a direct link.")


if __name__ == "__main__":
    create_sample_courses_excel()
